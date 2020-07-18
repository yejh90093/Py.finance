import datetime
import json

import numpy
import requests
from io import StringIO
import pandas as pd
import numpy as np


def readAll():
    currentDate = datetime.datetime.utcnow()
    condition = True
    while condition:
        # loop body here
        currentDate = currentDate - datetime.timedelta(days=1)
        currentDateStr = currentDate.strftime("%Y%m%d")
        r = requests.post(
            'http://www.twse.com.tw/exchangeReport/MI_INDEX?response=csv&date=' + currentDateStr + '&type=ALL')
        if r.text != "":
            condition = False

    str_list = []
    for i in r.text.split('\n'):
        if len(i.split('",')) == 17 and i[0] != '=':
            i = i.strip(",\r\n")
            str_list.append(i)

    df = pd.read_csv(StringIO("\n".join(str_list)))
    pd.set_option('display.max_rows', None)
    return df


# Valid intervals: [1m, 2m, 5m, 15m, 30m, 60m, 90m, 1h, 1d, 5d, 1wk, 1mo, 3mo]
def getFinanceData(id, start, end, interval):
    #id = "4142"
    site = "https://query1.finance.yahoo.com/v7/finance/download/" + id + ".TW?" \
                                                                          "period1=" + end + \
           "&period2=" + start + \
           "&interval=" + interval + \
           "&events=history" \
           "&crumb=hP2rOschxO0"
    response = requests.get(site)
    #print(site)
    return response

# Valid intervals: [1m, 2m, 5m, 15m, 30m, 60m, 90m, 1h, 1d, 5d, 1wk, 1mo, 3mo]
def getFinanceChart(id, start, end, interval):
    #id = "4142"
    site = "https://query1.finance.yahoo.com/v8/finance/chart/" + id + ".TW?" \
           "symbol=" + id + ".TW" + \
           "&period1=" + end + \
           "&period2=" + start + \
           "&interval=" + interval + \
           "&includePrePost=true" \
           "&events=div%2Csplit"


    response = requests.get(site)
    #print(site)
    return response

# https://query1.finance.yahoo.com/v8/finance/chart/2330.TW?
# symbol=2330.TW&period1=1594000000&period2=1594460555&interval=1m&includePrePost=true&events=div%2Csplit

def dataTextToArray(text):
    df = pd.read_csv(StringIO(text))

    for index, row in df.iterrows():
        if row['Open'] is None:
            df.at[index, 'Open'] = df.iloc[index-1]['Open']
            df.at[index, 'High'] = df.iloc[index-1]['High']
            df.at[index, 'Low'] = df.iloc[index-1]['Low']
            df.at[index, 'Close'] = df.iloc[index-1]['Close']

    return df.values

def dataJsonToArray(text):
    jsonData = json.loads(text)

    arrTimestamp = jsonData['chart']['result'][0]['timestamp']
    arrOpen = jsonData['chart']['result'][0]['indicators']['quote'][0]['open']
    arrHigh = jsonData['chart']['result'][0]['indicators']['quote'][0]['high']
    arrLow = jsonData['chart']['result'][0]['indicators']['quote'][0]['low']
    arrClose = jsonData['chart']['result'][0]['indicators']['quote'][0]['close']
    df = pd.DataFrame(np.vstack((arrTimestamp, arrOpen, arrHigh, arrLow, arrClose)).T)

    tradingPost = jsonData['chart']['result'][0]['meta']['tradingPeriods']['post']
    # print(tradingPost)

    #finalTimestamp = tradingPost[len(tradingPost)-1][0]['start']
    #finalClose = jsonData['chart']['result'][0]['meta']['regularMarketPrice']
    #finalDf = pd.DataFrame([finalTimestamp, finalClose, finalClose, finalClose, finalClose])

    # print(finalTimestamp)

    #df = df.append(finalDf.T, ignore_index=True)
    df.columns = ['timestamp', 'open', 'high', 'low', 'close']

    for index, row in df.iterrows():
        df.at[index, 'timestamp'] = datetime.datetime.fromtimestamp(int(row['timestamp']))
        #if row['open'] is None:
        #    df.at[index, 'open'] = df.iloc[index-1]['open']
        #    df.at[index, 'high'] = df.iloc[index-1]['high']
        #    df.at[index, 'low'] = df.iloc[index-1]['low']
        #    df.at[index, 'close'] = df.iloc[index-1]['close']

    # print(df)
    return df.dropna().values


def arrayWilliamsR(dataArray, nDay):
    arrWilliamsR = []
    NH = []
    NL = []
    n = 0
    nDay = nDay - 1

    for day in dataArray:

        if n >= nDay:
            NH = []
            NL = []
            for i in range(n, n - nDay, -1):
                NH.append(dataArray[i][2])
                NL.append(dataArray[i][3])

        if len(NH) == 0:
            B = dataArray[n]

            A = numpy.array([None, None, None])
            B = numpy.append(B, A)
        else:
            B = dataArray[n]
            W = (max(NH) - dataArray[n][4]) / (max(NH) - min(NL)) * 100

            A = numpy.array([round(max(NH),2), round(min(NL),2), round(W,2)])

            B = numpy.append(B, A)

        # print (n, dataArray[n][0], A)
        arrWilliamsR.append(B)
        n = n + 1

    return arrWilliamsR


def arrayKD(dataArray, nRange):
    arrKD = []
    lastK = 0.5
    lastD = 0.5
    n = 0

    for dataRow in dataArray:
        RSV = 0
        K = 0
        D = 0

        if n >= nRange:
            nHigh = dataArray[n][2]
            nLow = dataArray[n][3]

            for i in range(n, n - nRange, -1):
                nHigh = dataArray[i][2] if dataArray[i][2] > nHigh else nHigh
                nLow = dataArray[i][3] if dataArray[i][3] < nLow else nLow

            RSV = (dataArray[n][4] - nLow) / (nHigh - nLow) if (nHigh - nLow) > 0 else 1
            K = (1/3) * RSV + (2/3) * lastK
            D = (1/3) * K + (2/3) * lastD

            lastK = K
            lastD = D

            A = numpy.array([round(RSV*100, 2), round(K*100, 2), round(D*100, 2)])
            B = dataArray[n]
            B = numpy.append(B, A)
            arrKD.append(B)
        else:
            B = dataArray[n]
            A = numpy.array([None, None, None])
            B = numpy.append(B, A)
            arrKD.append(B)

        #print (n, dataArray[n][0], A)
        n = n + 1

    return arrKD



def arrayRSI(dataArray, nDay):
    arrRSI = []
    n = 0
    nDay = nDay - 1

    lastPlus = None
    lastMinus = None

    for day in dataArray:
        sumPlus = 0
        sumMinus = 0

        if n >= nDay:

            if n == nDay:
                for i in range(n, n - nDay, -1):
                    dayDiff = dataArray[i][4] - dataArray[i - 1][4]
                    if numpy.math.isnan(dayDiff): dayDiff = 0

                    if dayDiff < 0:
                        sumMinus = sumMinus + abs(dayDiff)
                    else:
                        sumPlus = sumPlus + abs(dayDiff)
            else:
                dayDiff = dataArray[n][4] - dataArray[n - 1][4]
                if numpy.math.isnan(dayDiff): dayDiff = 0

                if dayDiff < 0:
                    sumPlus = (lastPlus * nDay / (nDay + 1))
                    sumMinus = (lastMinus * nDay / (nDay + 1)) + (abs(dayDiff) / (nDay + 1))
                else:
                    sumPlus = (lastPlus * nDay / (nDay + 1)) + (abs(dayDiff) / (nDay + 1))
                    sumMinus = (lastMinus * nDay / (nDay + 1))

            lastPlus = sumPlus
            lastMinus = sumMinus

            avgPlus = sumPlus / (nDay + 1)
            avgMinus = sumMinus / (nDay + 1)

            if avgPlus + avgMinus == 0:
                dayRSI = None
                A = numpy.array([None])
            else:
                dayRSI = avgPlus / (avgPlus + avgMinus)
                A = numpy.array([round(dayRSI * 100, 2)])

            B = dataArray[n]
            B = numpy.append(B, A)
            arrRSI.append(B)
        else:
            B = dataArray[n]
            A = numpy.array([None])
            B = numpy.append(B, A)
            arrRSI.append(B)

        # print (n, dataArray[n][0], A)
        n = n + 1

    return arrRSI


def arrayDMI(dataArray, nDay):
    arrDMI = []
    n = 0
    # nDay = nDay - 1
    for day in dataArray:
        plusDM = 0
        minusDM = 0
        plusDI = 0
        minusDI = 0

        Ht_Lt = 0
        Ht_Ct_1 = 0
        Lt_Ct_1 = 0
        TR = 0

        if n >= nDay:

            plusDM = dataArray[n][2] - dataArray[n - 1][2]
            minusDM = dataArray[n - 1][3] - dataArray[n][3]

            Ht_Lt = dataArray[n][2] - dataArray[n][3]
            Ht_Ct_1 = dataArray[n][2] - dataArray[n - 1][4]
            Lt_Ct_1 = dataArray[n][3] - dataArray[n - 1][4]
            Lt_Ct_1 = Lt_Ct_1 if Lt_Ct_1 > 0 else 0

            TR = max(Ht_Lt, Ht_Ct_1, Lt_Ct_1)

            if TR == 0:
                plusDI = 0
                minusDI = 0
            else:
                plusDI = plusDM / TR * 100
                minusDI = minusDM / TR * 100

            # Normalize DI
            plusDI = plusDI if plusDI > 0 else 0
            minusDI = minusDI if minusDI > 0 else 0
            # Normalize DI

            A = numpy.array([round(plusDI, 2), round(minusDI, 2)])
            B = dataArray[n]
            B = numpy.append(B, A)
            arrDMI.append(B)
        else:
            B = dataArray[n]
            A = numpy.array([None, None])
            B = numpy.append(B, A)
            arrDMI.append(B)

        # print (n, dataArray[n][0], A)
        n = n + 1

    return arrDMI


def arrayMTM(dataArray, nMTM, nMA):
    arrMTM = []
    n = 0
    MA = None
    MTM_MA_result = None

    for day in dataArray:
        if n >= nMTM:
            MTM = dataArray[n][4] - dataArray[n - nMTM][4]

            if n >= nMTM + nMA:
                sum_MA = 0
                MTM_direct = dataArray[n-1][4] - dataArray[n - nMTM -1][4]

                for i in range(n, n - nMA, -1):
                    sum_MA += dataArray[i][4] - dataArray[i - nMTM][4]
                MA = sum_MA / nMA
                MTM_MA_result = "●" if (MTM > MA and MTM > 0 and MTM > MTM_direct) else ""

            A = numpy.array([MTM, MA, MTM_MA_result])
            B = dataArray[n]
            B = numpy.append(B, A)
            arrMTM.append(B)
        else:
            B = dataArray[n]
            A = numpy.array([None, None, None])
            B = numpy.append(B, A)
            arrMTM.append(B)

        #print (n, dataArray[n][0], A)
        n = n + 1

    return arrMTM


def arrayMACD(dataArray, ema1, ema2, diff):

    arrMACD = []
    n = 0
    last_ema1 = 0
    curr_ema1 = 0
    last_ema2 = 0

    last_diff = 0
    curr_diff = 0
    direction = ""

    for data in dataArray:

        if n >= ema1:

            if n == ema1:
                #print (dataArray[0:ema1, 4])
                curr_ema1 = np.nansum(dataArray[0:ema1, 4])/ema1
                last_ema1 = curr_ema1
            else:
                curr_ema1 = (1-(2/(ema1+1))) * last_ema1 + (2/(ema1+1)) * (dataArray[n-1][4])
                last_ema1 = curr_ema1

            if n >= ema2:
                if n == ema2:
                    #print(dataArray[0:ema2, 4])
                    curr_ema2 = np.nansum(dataArray[0:ema2, 4])/ema2
                    last_ema2 = curr_ema2
                else:
                    curr_ema2 = (1-(2/(ema2+1))) * last_ema2 + (2/(ema2+1)) * (dataArray[n-1][4])
                    last_ema2 = curr_ema2

                curr_diff = curr_ema1 - curr_ema2
                direction = "↗" if (curr_diff > last_diff) else "↘"
                direction = "=" if (curr_diff == last_diff) else direction
                last_diff = curr_diff
                A = numpy.array([round(curr_ema1, 2), round(curr_ema2, 2), round(curr_diff, 2), direction])
            else:
                A = numpy.array([round(curr_ema1, 2), None, None, None])

            B = dataArray[n]
            B = numpy.append(B, A)
            arrMACD.append(B)
        else:
            B = dataArray[n]
            A = numpy.array([None, None, None, None])
            B = numpy.append(B, A)
            arrMACD.append(B)

        # print (n, dataArray[n], A)
        n = n + 1

    return arrMACD

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()


